import tkinter as tk
import openpyxl


class ProcessData():
    new_rows =[]

    def __init__(self,path,treeView):
        self.path = path
        self.treeView = treeView

    def load_data(self):

        for item in self.treeView.get_children():
            self.treeView.delete(item)
        
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active
 
        list_values = list(sheet.values)

        for col_name in list_values[0]:
            self.treeView.heading(col_name, text=col_name)

        for value_tuple in list_values[1:]:
            self.treeView.insert('',tk.END,values=value_tuple)
    
    def insert_row(self,name,plate,color,make,year,serviced):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active
        new_values = (name,plate,color,make,year,serviced)
       
        sheet.append(new_values)
        workbook.save(self.path)
        self.treeView.insert('',tk.END,values=new_values)

  
    def plate_exists(self,plate):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active
        list_values = list(sheet.values)

        for item in list_values:
            if item[1].strip() == plate and plate != "Plate":
                return True
        
        return False

    def filter_data(self, name, plate, color, make, year, serviced):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active
        list_values = list(sheet.values)

        for item in self.treeView.get_children():
            # Get the values of the current row from the treeview item
            row_values = self.treeView.item(item)['values']
        
            row_name = row_values[0].strip() if row_values[0] else ""
            row_plate = row_values[1].strip() if row_values[1] else ""
            row_color = row_values[2].strip() if row_values[2] else ""
            row_make = row_values[3].strip() if row_values[3] else ""
            row_year = str(row_values[4]).strip() if row_values[4] else ""
            row_serviced = row_values[5].strip() if row_values[5] else ""

            matches = True
            if name and name != "Name" and name != row_name:
                matches = False
            if plate and plate != "Plate" and plate != row_plate:
                matches = False
            if color and color != "Color" and color != row_color:
                matches = False
            if make and make != "Make" and make != row_make:
                matches = False
            if year and year != "Year" and str(year) != row_year:
                matches = False
            if serviced and serviced != row_serviced:
                matches = False

            # If not matching, delete the row from the Treeview
            if not matches:
                self.treeView.delete(item)

        

        

