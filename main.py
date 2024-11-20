import tkinter as tk
from tkinter import ttk

import openpyxl

class ProcessData():
    def __init__(self,path,treeView):
        self.path = path
        self.treeView = treeView

    def load_data(self):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active

        list_values = list(sheet.values)

        for col_name in list_values[0]:
            self.treeView.heading(col_name, text=col_name)

        for value_tuple in list_values[1:]:
            self.treeView.insert('',tk.END,values=value_tuple)
    
    def insert_row(self,name,plate,color,make,year,serviced):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active
        new_values = (name,plate,color,make,year,serviced)
       
        sheet.append(new_values)
        workbook.save(self.path)

        self.treeView.insert('',tk.END,values=new_values)

    def filter_data(self, name, plate, color, make, year, serviced):
        for item in self.treeView.get_children():
            self.treeView.delete(item)
            
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active
        list_values = list(sheet.values)

        for value_tuple in list_values[1:]:
            row_name = value_tuple[0].strip() if value_tuple[0] else ""
            row_plate = value_tuple[1].strip() if value_tuple[1] else ""
            row_color = value_tuple[2].strip() if value_tuple[2] else ""
            row_make = value_tuple[3].strip() if value_tuple[3] else ""
            row_year = str(value_tuple[4]).strip() if value_tuple[4] else ""
            row_serviced = value_tuple[5].strip() if value_tuple[5] else ""

           
            matches = True
            if name and name != "Name" and name != row_name:
                matches = False
            if plate and plate != "Plate" and plate != row_plate:
                matches = False
            if color and color != "Color" and color != row_color:
                matches = False
            if make and make != "Make" and make != row_make:
                matches = False
            if year and year !="Year" and str(year) != row_year:
                matches = False
            if serviced and serviced != row_serviced:
                matches = False

            if matches:
                self.treeView.insert('', tk.END, values=value_tuple)



root = tk.Tk()

style = ttk.Style(root)

root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-light")

colors = ["White","Green","Black","Blue","Gray","Red"]
makes = ["Ford","Audi","BMW","Mercedes-benz","Dodge","Toyota","Honda","Hyundai","Mazda"]

def mode_switch():
    if mode_butt.instate(["selected"]):
        style.theme_use("forest-dark")
    else:
        style.theme_use("forest-light")


frame = ttk.Frame(root)
frame.pack()


label_frame = ttk.LabelFrame(frame,text="Insert Row")
label_frame.grid(row=0,column=0,padx=5,pady=10)

name_entry_widget=ttk.Entry(label_frame)
name_entry_widget.insert("0","Name")
name_entry_widget.bind("<FocusIn>",lambda n:name_entry_widget.delete("0","end"))
name_entry_widget.grid(row=0,column=0,padx=5,pady=(0,5),sticky="ew")

plate_entry_widget=ttk.Entry(label_frame)
plate_entry_widget.insert("0","Plate")
plate_entry_widget.bind("<FocusIn>",lambda p:plate_entry_widget.delete("0","end"))
plate_entry_widget.grid(row=1,column=0,padx=5,pady=(0,5),sticky="ew")

year_spin_widget = ttk.Spinbox(label_frame,from_=1970,to=2025)
year_spin_widget.insert("0","Year")
year_spin_widget.grid(row=2,column=0,padx=5,pady=(0,5))

combo_widget_make =ttk.Combobox(label_frame,values=makes)
combo_widget_make.insert("0","Make")
combo_widget_make.grid(row=3,column=0,padx=5,pady=(0,5),sticky="ew")

combo_widget_colors = ttk.Combobox(label_frame,values=colors)
combo_widget_colors.insert("0","Color")
combo_widget_colors.grid(row=4,column=0,padx=5,pady=(0,5),sticky="ew")

service_butt = tk.BooleanVar()
check_button =ttk.Checkbutton(label_frame,text="Auto Service",variable=service_butt)
check_button.grid(row=5,column=0,padx=5,pady=(0,5),sticky="ew")


separator = ttk.Separator(label_frame)
separator.grid(row=7,column=0,padx=5,pady=(10,5),sticky="ew")

mode_butt = ttk.Checkbutton(label_frame,text="Lite/Dark Mode",style="Switch",command=mode_switch)
mode_butt.grid(row=9,column=0,padx=5,pady=(0,10),sticky="nsew")


treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0,column=1,padx=10)

cols = ("Name","Plate","Color","Make","Year","Auto Service Status")


scroll_bar = ttk.Scrollbar(treeFrame)
scroll_bar.pack(side="right",fill="y")



treeView = ttk.Treeview(treeFrame,show="headings",yscrollcommand=scroll_bar.set, columns=cols,height=15) 
treeView.column("Name",width=100)
treeView.column("Plate",width=70)
treeView.column("Color",width=60)
treeView.column("Make",width=100)
treeView.column("Year",width=50)
treeView.column("Auto Service Status",width=120)


treeView.pack()

scroll_bar.config(command=treeView.yview)

path = r"C:\Users\maryb\OneDrive\Desktop\python exel app\car_owners.xlsx"
car_owners = ProcessData(path,treeView)

car_owners.load_data()

def reset_widgets():
    name_entry_widget.delete(0,"end")
    name_entry_widget.insert(0,"Name")
    plate_entry_widget.delete(0,"end")
    plate_entry_widget.insert(0,"Plate")
    combo_widget_colors.delete(0,"end")
    combo_widget_colors.insert(0,"Color")
    combo_widget_make.delete(0,"end")
    combo_widget_make.insert(0,"Make")
    year_spin_widget.delete(0,"end")
    year_spin_widget.insert(0,"Year")
    check_button.state(["!selected"])  

def search_data():
    name = name_entry_widget.get().strip()  # Trim whitespace
    plate = plate_entry_widget.get().strip()
    color = combo_widget_colors.get().strip()
    make = combo_widget_make.get().strip()
    year = year_spin_widget.get().strip() 
    serviced = "Serviced" if service_butt.get() else "Not serviced"

    car_owners.filter_data(name, plate, color, make, year, serviced)

def insert_data():
    try:
        name = name_entry_widget.get()
        plate = plate_entry_widget.get()
        color = combo_widget_colors.get()
        make = combo_widget_make.get()
        year = int(year_spin_widget.get())
        serviced = "Serviced" if service_butt.get() else "Not serviced"
    except Exception as e:
        reset_widgets()
    
    car_owners.insert_row(name,plate,color,make,year,serviced)
    
    reset_widgets()


reset_buttn = ttk.Button(label_frame,text="Reset",command=car_owners.load_data)
reset_buttn.grid(row=8,column=0,padx=5,pady=(5,5),sticky="ew")

filter_buttn = ttk.Button(label_frame,text="Filter",command=search_data)
filter_buttn.grid(row=6,column=0,padx=5,pady=(0,5),sticky="ew")


insert_buttn = ttk.Button(label_frame,text="Insert",command=insert_data)
insert_buttn.grid(row=7,column=0,padx=5,pady=(5,5),sticky="ew")



root.mainloop()